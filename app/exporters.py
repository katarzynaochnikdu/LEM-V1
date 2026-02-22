"""
Eksport wynikow LEM do formatow czytelnych dla uzytkownika.
"""

from __future__ import annotations

from abc import ABC, abstractmethod
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


COMPETENCY_LABELS: Dict[str, str] = {
    "delegowanie": "Delegowanie",
    "decyzje": "Podejmowanie decyzji",
    "priorytety": "Priorytety",
    "feedback": "Feedback",
    "podejmowanie_decyzji": "Podejmowanie decyzji",
    "okreslanie_priorytetow": "Priorytety",
    "udzielanie_feedbacku": "Feedback",
}

EXPORT_FILE_TYPES: Dict[str, Tuple[str, str]] = {
    "json": ("application/json; charset=utf-8", "json"),
    "txt": ("text/plain; charset=utf-8", "txt"),
    "html": ("text/html; charset=utf-8", "html"),
    "excel": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
    "pdf_full": ("application/pdf", "pdf"),
    "pdf_summary": ("application/pdf", "pdf"),
}


class BaseExporter(ABC):
    @abstractmethod
    def export(self, payload: Dict[str, Any]) -> bytes:
        raise NotImplementedError

    def _normalize(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        participant_id = str(payload.get("participant_id") or "session")
        generated_at = payload.get("generated_at") or datetime.utcnow().isoformat()
        results = payload.get("results") or {}
        selected = payload.get("selected_competencies") or list(results.keys())
        response_text = str(payload.get("response_text") or "")
        return {
            "participant_id": participant_id,
            "generated_at": generated_at,
            "results": results,
            "selected_competencies": selected,
            "response_text": response_text,
        }

    def _competency_name(self, competency_id: str) -> str:
        return COMPETENCY_LABELS.get(competency_id, competency_id.replace("_", " ").title())

    def _iter_competencies(self, payload: Dict[str, Any]) -> List[Tuple[str, Dict[str, Any]]]:
        normalized = self._normalize(payload)
        entries: List[Tuple[str, Dict[str, Any]]] = []
        for comp_id in normalized["selected_competencies"]:
            result = normalized["results"].get(comp_id)
            if result:
                entries.append((comp_id, result))
        if not entries:
            for comp_id, result in normalized["results"].items():
                entries.append((comp_id, result))
        return entries


class TXTExporter(BaseExporter):
    def export(self, payload: Dict[str, Any]) -> bytes:
        data = self._normalize(payload)
        lines: List[str] = []
        lines.append("=" * 60)
        lines.append("RAPORT OCENY KOMPETENCJI LEM")
        lines.append("=" * 60)
        lines.append(f"Uczestnik: {data['participant_id']}")
        lines.append(f"Data raportu: {data['generated_at']}")
        lines.append("")

        for index, (comp_id, result) in enumerate(self._iter_competencies(data), start=1):
            scored = result.get("scored") or {}
            feedback = result.get("feedback") or {}
            mapped = result.get("mapped") or {}
            lines.append("-" * 60)
            lines.append(f"{index}. {self._competency_name(comp_id)}")
            lines.append("-" * 60)
            lines.append(
                f"Wynik: {scored.get('overallScore', 0):.2f}/4.00 | Poziom: {scored.get('levelName', 'Brak danych')}"
            )
            lines.append(
                f"Pokrycie wymiarow: {mapped.get('detectedCount', 0)}/{mapped.get('totalCount', 0)}"
            )
            lines.append("")
            lines.append("Podsumowanie:")
            lines.append(str(feedback.get("summary") or "Brak"))
            lines.append("")
            lines.append("Rekomendacja:")
            lines.append(str(feedback.get("recommendation") or "Brak"))
            lines.append("")

            strengths = feedback.get("strengths") or []
            lines.append("Mocne strony:")
            if strengths:
                for item in strengths:
                    lines.append(f"- {item}")
            else:
                lines.append("- Brak")
            lines.append("")

            development = feedback.get("developmentAreas") or []
            lines.append("Obszary do rozwoju:")
            if development:
                for item in development:
                    lines.append(f"- {item}")
            else:
                lines.append("- Brak")
            lines.append("")

            lines.append("Oceny wymiarow:")
            dimensions = (scored.get("dimensions") or [])
            if dimensions:
                for dim in dimensions:
                    lines.append(
                        f"- {dim.get('name', 'wymiar')}: {float(dim.get('score', 0)):.2f}/1.00 "
                        f"(waga {float(dim.get('weight', 0)):.2f}, punkty {float(dim.get('points', 0)):.2f})"
                    )
            else:
                lines.append("- Brak danych")
            lines.append("")

            mapped_dims = (mapped.get("dimensions") or [])
            lines.append("Przykladowe dowody:")
            has_evidence = False
            for dim in mapped_dims:
                evidence = str(dim.get("evidence") or "").strip()
                if evidence:
                    has_evidence = True
                    lines.append(f"- {dim.get('name', 'wymiar')}: {evidence}")
            if not has_evidence:
                lines.append("- Brak")
            lines.append("")

        if data["response_text"]:
            lines.append("-" * 60)
            lines.append("Analizowana odpowiedz (skrot)")
            lines.append("-" * 60)
            text = data["response_text"].strip()
            lines.append(text[:1200] + ("..." if len(text) > 1200 else ""))
            lines.append("")

        return "\n".join(lines).encode("utf-8")


class HTMLExporter(BaseExporter):
    def export(self, payload: Dict[str, Any]) -> bytes:
        data = self._normalize(payload)
        cards: List[str] = []
        for comp_id, result in self._iter_competencies(data):
            scored = result.get("scored") or {}
            feedback = result.get("feedback") or {}
            mapped = result.get("mapped") or {}
            dim_rows = []
            for dim in scored.get("dimensions") or []:
                dim_rows.append(
                    "<tr>"
                    f"<td>{_escape(str(dim.get('name', '')))}</td>"
                    f"<td>{float(dim.get('score', 0)):.2f}/1.00</td>"
                    f"<td>{float(dim.get('weight', 0)):.2f}</td>"
                    f"<td>{float(dim.get('points', 0)):.2f}</td>"
                    f"<td>{_escape(str(dim.get('rationale') or ''))}</td>"
                    "</tr>"
                )
            evidence_list = []
            for dim in mapped.get("dimensions") or []:
                ev = str(dim.get("evidence") or "").strip()
                if ev:
                    evidence_list.append(
                        f"<li><strong>{_escape(str(dim.get('name', '')))}:</strong> {_escape(ev)}</li>"
                    )

            strengths = "".join(f"<li>{_escape(str(x))}</li>" for x in (feedback.get("strengths") or [])) or "<li>Brak</li>"
            development = "".join(f"<li>{_escape(str(x))}</li>" for x in (feedback.get("developmentAreas") or [])) or "<li>Brak</li>"

            cards.append(
                f"""
                <section class="card">
                    <h2>{_escape(self._competency_name(comp_id))}</h2>
                    <p class="score">Wynik: <strong>{float(scored.get('overallScore', 0)):.2f}/4.00</strong> | Poziom: {_escape(str(scored.get('levelName', 'Brak danych')))}</p>
                    <p>Pokrycie wymiarow: {mapped.get('detectedCount', 0)}/{mapped.get('totalCount', 0)}</p>
                    <h3>Podsumowanie</h3>
                    <p>{_escape(str(feedback.get("summary") or "Brak"))}</p>
                    <h3>Rekomendacja</h3>
                    <p>{_escape(str(feedback.get("recommendation") or "Brak"))}</p>
                    <div class="cols">
                        <div>
                            <h3>Mocne strony</h3>
                            <ul>{strengths}</ul>
                        </div>
                        <div>
                            <h3>Obszary do rozwoju</h3>
                            <ul>{development}</ul>
                        </div>
                    </div>
                    <h3>Oceny wymiarow</h3>
                    <table>
                        <thead><tr><th>Wymiar</th><th>Ocena</th><th>Waga</th><th>Punkty</th><th>Uzasadnienie</th></tr></thead>
                        <tbody>{''.join(dim_rows) if dim_rows else '<tr><td colspan="5">Brak danych</td></tr>'}</tbody>
                    </table>
                    <h3>Dowody</h3>
                    <ul>{''.join(evidence_list) if evidence_list else '<li>Brak</li>'}</ul>
                </section>
                """
            )

        html = f"""
        <!doctype html>
        <html lang="pl">
        <head>
            <meta charset="utf-8" />
            <meta name="viewport" content="width=device-width, initial-scale=1" />
            <title>Raport LEM - { _escape(data['participant_id']) }</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 24px; color: #1f2937; background: #f9fafb; }}
                .header {{ background: #111827; color: #fff; padding: 16px 20px; border-radius: 8px; margin-bottom: 16px; }}
                .header p {{ margin: 4px 0; opacity: .92; }}
                .card {{ background: white; border: 1px solid #e5e7eb; border-radius: 8px; padding: 16px; margin-bottom: 14px; }}
                .score {{ font-size: 16px; }}
                .cols {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
                h1, h2, h3 {{ margin: 0 0 8px 0; }}
                h2 {{ margin-top: 8px; }}
                ul {{ margin: 0; padding-left: 18px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
                th, td {{ border: 1px solid #e5e7eb; padding: 6px; text-align: left; font-size: 13px; vertical-align: top; }}
                th {{ background: #f3f4f6; }}
                @media print {{
                    body {{ margin: 10mm; background: white; }}
                    .card {{ break-inside: avoid; }}
                }}
            </style>
        </head>
        <body>
            <header class="header">
                <h1>Raport oceny kompetencji LEM</h1>
                <p>Uczestnik: {_escape(data['participant_id'])}</p>
                <p>Data raportu: {_escape(data['generated_at'])}</p>
            </header>
            {''.join(cards) if cards else '<p>Brak wynikow do eksportu.</p>'}
        </body>
        </html>
        """
        return html.encode("utf-8")


class ExcelExporter(BaseExporter):
    def export(self, payload: Dict[str, Any]) -> bytes:
        data = self._normalize(payload)
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Podsumowanie"
        ws_summary.append(["Uczestnik", "Data raportu", "Kompetencja", "Wynik", "Poziom", "Wykryte wymiary", "Laczna liczba wymiarow"])
        _bold_row(ws_summary, 1)

        ws_dimensions = wb.create_sheet("Wymiary")
        ws_dimensions.append(["Kompetencja", "Wymiar", "Ocena", "Waga", "Punkty", "Uzasadnienie"])
        _bold_row(ws_dimensions, 1)

        ws_evidence = wb.create_sheet("Dowody")
        ws_evidence.append(["Kompetencja", "Wymiar", "Czy obecny", "Dowody", "Notatki"])
        _bold_row(ws_evidence, 1)

        for comp_id, result in self._iter_competencies(data):
            scored = result.get("scored") or {}
            mapped = result.get("mapped") or {}
            cname = self._competency_name(comp_id)

            ws_summary.append(
                [
                    data["participant_id"],
                    data["generated_at"],
                    cname,
                    float(scored.get("overallScore", 0)),
                    str(scored.get("levelName") or ""),
                    int(mapped.get("detectedCount", 0)),
                    int(mapped.get("totalCount", 0)),
                ]
            )

            for dim in scored.get("dimensions") or []:
                ws_dimensions.append(
                    [
                        cname,
                        dim.get("name", ""),
                        float(dim.get("score", 0)),
                        float(dim.get("weight", 0)),
                        float(dim.get("points", 0)),
                        str(dim.get("rationale") or ""),
                    ]
                )

            for dim in mapped.get("dimensions") or []:
                ws_evidence.append(
                    [
                        cname,
                        dim.get("name", ""),
                        bool(dim.get("present", False)),
                        str(dim.get("evidence") or ""),
                        str(dim.get("notes") or ""),
                    ]
                )

        _autosize_columns(ws_summary)
        _autosize_columns(ws_dimensions)
        _autosize_columns(ws_evidence)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()


class PDFExporter(BaseExporter):
    def __init__(self, summary_only: bool = False):
        self.summary_only = summary_only

    def export(self, payload: Dict[str, Any]) -> bytes:
        data = self._normalize(payload)
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=1.8 * cm,
            rightMargin=1.8 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            title="Raport LEM",
        )
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="H1LEM", parent=styles["Heading1"], fontSize=17, spaceAfter=8))
        styles.add(ParagraphStyle(name="H2LEM", parent=styles["Heading2"], fontSize=13, spaceBefore=10, spaceAfter=6))
        styles.add(ParagraphStyle(name="BodyLEM", parent=styles["BodyText"], fontSize=10, leading=14))

        story: List[Any] = []
        story.append(Paragraph("Raport oceny kompetencji LEM", styles["H1LEM"]))
        story.append(Paragraph(f"Uczestnik: {data['participant_id']}", styles["BodyLEM"]))
        story.append(Paragraph(f"Data raportu: {data['generated_at']}", styles["BodyLEM"]))
        story.append(Spacer(1, 8))

        for comp_id, result in self._iter_competencies(data):
            scored = result.get("scored") or {}
            feedback = result.get("feedback") or {}
            mapped = result.get("mapped") or {}
            story.append(Paragraph(self._competency_name(comp_id), styles["H2LEM"]))
            story.append(
                Paragraph(
                    f"Wynik: <b>{float(scored.get('overallScore', 0)):.2f}/4.00</b> | "
                    f"Poziom: {str(scored.get('levelName') or 'Brak danych')}",
                    styles["BodyLEM"],
                )
            )
            story.append(
                Paragraph(
                    f"Pokrycie wymiarow: {int(mapped.get('detectedCount', 0))}/{int(mapped.get('totalCount', 0))}",
                    styles["BodyLEM"],
                )
            )
            story.append(Spacer(1, 4))
            story.append(Paragraph(f"<b>Podsumowanie:</b> {str(feedback.get('summary') or 'Brak')}", styles["BodyLEM"]))
            story.append(Paragraph(f"<b>Rekomendacja:</b> {str(feedback.get('recommendation') or 'Brak')}", styles["BodyLEM"]))
            story.append(Spacer(1, 4))

            if not self.summary_only:
                rows = [["Wymiar", "Ocena", "Waga", "Punkty"]]
                for dim in scored.get("dimensions") or []:
                    rows.append(
                        [
                            str(dim.get("name") or ""),
                            f"{float(dim.get('score', 0)):.2f}/1.00",
                            f"{float(dim.get('weight', 0)):.2f}",
                            f"{float(dim.get('points', 0)):.2f}",
                        ]
                    )
                if len(rows) == 1:
                    rows.append(["Brak danych", "-", "-", "-"])

                table = Table(rows, colWidths=[8.2 * cm, 2.5 * cm, 2.2 * cm, 2.2 * cm])
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(Spacer(1, 2))
                story.append(table)
                story.append(Spacer(1, 6))

        doc.build(story)
        return buf.getvalue()


def export_report(export_format: str, payload: Dict[str, Any]) -> bytes:
    if export_format == "json":
        import json

        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    if export_format == "txt":
        return TXTExporter().export(payload)
    if export_format == "html":
        return HTMLExporter().export(payload)
    if export_format == "excel":
        return ExcelExporter().export(payload)
    if export_format == "pdf_full":
        return PDFExporter(summary_only=False).export(payload)
    if export_format == "pdf_summary":
        return PDFExporter(summary_only=True).export(payload)
    raise ValueError(f"Nieobslugiwany format eksportu: {export_format}")


def get_content_type(export_format: str) -> str:
    return EXPORT_FILE_TYPES.get(export_format, EXPORT_FILE_TYPES["json"])[0]


def get_filename(export_format: str, participant_id: str, timestamp: str | None = None) -> str:
    safe_participant = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in participant_id)[:40] or "session"
    ts = (timestamp or datetime.utcnow().isoformat()).replace(":", "-").replace(".", "-")
    ext = EXPORT_FILE_TYPES.get(export_format, EXPORT_FILE_TYPES["json"])[1]
    return f"lem-export-{safe_participant}-{ts}.{ext}"


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def _bold_row(ws, row_idx: int) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)


def _escape(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )
